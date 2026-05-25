"""
excel_export.py
---------------
Generates a styled in-memory xlsx workbook from the current tournament data.

Usage:
    from modules.excel_export import generate_workbook_bytes
    xlsx_bytes = generate_workbook_bytes()          # active location
    xlsx_bytes = generate_workbook_bytes("Indore")  # specific location

The returned bytes can be passed directly to st.download_button().
Excel files are NEVER written to disk — only produced on demand.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.excel_sync import HEADER_COLORS, SHEET_HEADERS, load_sheet, set_location, _active_location


# ---------------------------------------------------------------------------
# Internal helpers (openpyxl styling)
# ---------------------------------------------------------------------------

def _apply_header_style(ws, color_hex: str) -> None:
    fill  = PatternFill("solid", fgColor=color_hex)
    font  = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align
    ws.row_dimensions[1].height = 22


def _auto_column_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 14)


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    headers = SHEET_HEADERS.get(sheet_name, df.columns.tolist())

    if sheet_name in wb.sheetnames:
        position = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
    else:
        position = len(wb.sheetnames)

    ws = wb.create_sheet(sheet_name, position)
    ws.append(headers)
    _apply_header_style(ws, HEADER_COLORS.get(sheet_name, "4472C4"))

    for _, row in df.iterrows():
        ws.append([
            (bool(row[col]) if col in ("is_eliminated",) else row.get(col, None))
            for col in headers
        ])

    _auto_column_width(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_workbook_bytes(location: str | None = None) -> bytes:
    """Return the xlsx workbook for *location* (or the active location) as bytes.

    Loads all sheets from CSV, builds a styled openpyxl workbook in memory,
    and returns the raw bytes without writing anything to disk.
    """
    # Temporarily switch location if explicitly requested
    original = _active_location
    if location and location != original:
        set_location(location)

    try:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default blank sheet

        for sheet_name in SHEET_HEADERS:
            df = load_sheet(sheet_name)
            _write_sheet(wb, sheet_name, df)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        if location and location != original:
            set_location(original)

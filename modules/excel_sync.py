"""
excel_sync.py
-------------
Single source of truth for all Excel read/write operations.
All other modules call load_sheet() and save_sheet() — never touch openpyxl directly.

Workbook sheets:
  Players      — registered players with skill ratings
  Teams        — formed teams with win/loss counters
  Matches      — full match schedule and results
  MatchStats   — per-player per-match award flags
  Leaderboard  — auto-computed team standings (derived, overwritten on every save)
  PlayerStats  — auto-computed player award totals (derived, overwritten on every save)
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")

LOCATIONS = ["Bangalore", "Hyderabad", "Indore"]
_active_location: str = LOCATIONS[0]
EXCEL_PATH: str = os.path.join(DATA_DIR, f"tournament_{_active_location.lower()}.xlsx")


def set_location(loc: str) -> None:
    """Switch active location — updates EXCEL_PATH for all subsequent data calls."""
    global _active_location, EXCEL_PATH
    _active_location = loc
    EXCEL_PATH = os.path.join(DATA_DIR, f"tournament_{loc.lower()}.xlsx")

# ---------------------------------------------------------------------------
# Schema constants
# ---------------------------------------------------------------------------

AWARDS = [
    "queen_snatcher",
    "precision_player",
    "best_striker",
    "comeback_king",
]

SHEET_HEADERS = {
    "Players": ["player_id", "name", "skill_rating", "team_id"],
    "Teams": [
        "team_id", "team_name", "avg_skill",
        "wins", "losses", "is_eliminated",
    ],
    "Matches": [
        "match_id", "round", "team_a_id", "team_b_id",
        "winner_id", "loser_id", "bracket", "status", "date_played",
    ],
    "MatchStats": ["stat_id", "match_id", "player_id", "team_id"] + AWARDS,
    "Leaderboard": [
        "rank", "team_id", "team_name",
        "wins", "losses", "status", "total_awards",
    ],
    "PlayerStats": ["player_id", "name", "team_name"] + AWARDS + ["total_awards"],
}

HEADER_COLORS = {
    "Players":     "4472C4",   # Blue
    "Teams":       "70AD47",   # Green
    "Matches":     "ED7D31",   # Orange
    "MatchStats":  "7030A0",   # Purple
    "Leaderboard": "C00000",   # Red
    "PlayerStats": "00B0F0",   # Light blue
}

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _ensure_data_dir() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)


def _apply_header_style(ws, color_hex: str) -> None:
    fill = PatternFill("solid", fgColor=color_hex)
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 22


def _auto_column_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 14)


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    """Internal: recreate a sheet in `wb` from a DataFrame."""
    headers = SHEET_HEADERS.get(sheet_name, df.columns.tolist())

    # Remove and recreate the sheet at the same position
    if sheet_name in wb.sheetnames:
        position = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
    else:
        position = len(wb.sheetnames)

    ws = wb.create_sheet(sheet_name, position)
    ws.append(headers)
    _apply_header_style(ws, HEADER_COLORS.get(sheet_name, "4472C4"))

    for _, row in df.iterrows():
        ws.append([
            (bool(row[col]) if col in ("is_eliminated",) else row.get(col, None))
            for col in headers
        ])

    _auto_column_width(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def init_workbook() -> None:
    """Create the Excel workbook with all sheets and headers if it does not exist."""
    _ensure_data_dir()
    if os.path.exists(EXCEL_PATH):
        return

    wb = Workbook()
    wb.remove(wb.active)  # remove the default blank sheet

    for sheet_name, headers in SHEET_HEADERS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        _apply_header_style(ws, HEADER_COLORS[sheet_name])
        _auto_column_width(ws)
        ws.freeze_panes = "A2"

    wb.save(EXCEL_PATH)


def load_sheet(sheet_name: str) -> pd.DataFrame:
    """
    Load a sheet from the workbook into a DataFrame.
    Returns an empty DataFrame with the correct columns if the sheet is empty
    or the workbook does not exist yet.
    """
    _ensure_data_dir()
    if not os.path.exists(EXCEL_PATH):
        init_workbook()

    expected_cols = SHEET_HEADERS.get(sheet_name, [])
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, engine="openpyxl")
        if df.empty:
            return pd.DataFrame(columns=expected_cols)
        return df
    except Exception:
        return pd.DataFrame(columns=expected_cols)


def save_sheet(sheet_name: str, df: pd.DataFrame) -> None:
    """
    Overwrite a single named sheet in the workbook without affecting other sheets.
    Automatically calls update_derived_sheets() after saving source data.
    """
    _ensure_data_dir()
    if not os.path.exists(EXCEL_PATH):
        init_workbook()

    wb = load_workbook(EXCEL_PATH)
    _write_sheet(wb, sheet_name, df)
    wb.save(EXCEL_PATH)

    # Keep derived sheets in sync after every source-data write
    if sheet_name in ("Players", "Teams", "Matches", "MatchStats"):
        update_derived_sheets()


def get_next_id(sheet_name: str, id_column: str) -> int:
    """Return the next available integer ID for a sheet's ID column."""
    df = load_sheet(sheet_name)
    if df.empty or id_column not in df.columns or df[id_column].dropna().empty:
        return 1
    return int(df[id_column].dropna().max()) + 1


def update_derived_sheets() -> None:
    """
    Recompute Leaderboard and PlayerStats from source data and overwrite
    those sheets in the workbook.  Called automatically by save_sheet().
    """
    _ensure_data_dir()
    if not os.path.exists(EXCEL_PATH):
        return

    teams_df    = load_sheet("Teams")
    players_df  = load_sheet("Players")
    ms_df       = load_sheet("MatchStats")

    wb = load_workbook(EXCEL_PATH)

    # ---- Leaderboard -------------------------------------------------------
    if not teams_df.empty:
        lb = teams_df[["team_id", "team_name", "wins", "losses", "is_eliminated"]].copy()
        lb["wins"]   = pd.to_numeric(lb["wins"],   errors="coerce").fillna(0).astype(int)
        lb["losses"] = pd.to_numeric(lb["losses"], errors="coerce").fillna(0).astype(int)

        # Total awards per team
        if not ms_df.empty and not players_df.empty:
            pid_to_tid = players_df.dropna(subset=["player_id"]).set_index("player_id")["team_id"].to_dict()
            ms = ms_df.copy()
            ms["team_id"] = ms["player_id"].map(pid_to_tid)
            award_sum = (
                ms.groupby("team_id")[AWARDS]
                .sum()
                .sum(axis=1)
                .reset_index(name="total_awards")
            )
            lb = lb.merge(award_sum, on="team_id", how="left")
        else:
            lb["total_awards"] = 0

        lb["total_awards"] = lb["total_awards"].fillna(0).astype(int)
        lb["status"] = lb["is_eliminated"].apply(
            lambda x: "Eliminated" if (x is True or x == 1) else "Active"
        )
        lb = lb.sort_values(
            ["wins", "losses", "total_awards"],
            ascending=[False, True, False],
        ).reset_index(drop=True)
        lb.insert(0, "rank", range(1, len(lb) + 1))
        lb = lb[["rank", "team_id", "team_name", "wins", "losses", "status", "total_awards"]]
        _write_sheet(wb, "Leaderboard", lb)

    # ---- PlayerStats -------------------------------------------------------
    if not players_df.empty:
        ps = players_df[["player_id", "name", "team_id"]].copy()

        if not teams_df.empty:
            tname_map = teams_df.set_index("team_id")["team_name"].to_dict()
            ps["team_name"] = ps["team_id"].map(tname_map).fillna("Unassigned")
        else:
            ps["team_name"] = "Unassigned"

        if not ms_df.empty:
            award_totals = (
                ms_df.dropna(subset=["player_id"])
                .groupby("player_id")[AWARDS]
                .sum()
                .reset_index()
            )
            ps = ps.merge(award_totals, on="player_id", how="left")

        for a in AWARDS:
            if a not in ps.columns:
                ps[a] = 0
            ps[a] = pd.to_numeric(ps[a], errors="coerce").fillna(0).astype(int)

        ps["total_awards"] = ps[AWARDS].sum(axis=1)
        ps = ps[["player_id", "name", "team_name"] + AWARDS + ["total_awards"]]
        _write_sheet(wb, "PlayerStats", ps)

    wb.save(EXCEL_PATH)
